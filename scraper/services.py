"""
Beauty Salon Scraper via Google Maps Places API
Парсинг салонов красоты → Excel отчёт
"""

import time
import re
import requests
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from collections import Counter, defaultdict
from datetime import datetime

# ─── НАСТРОЙКИ ────────────────────────────────────────────────────────────────

API_KEY = "AIzaSyAJPy_Abo-DZYgvZD9LEv4YA7KcNmBZybo"   # ← вставьте свой ключ

# Регионы для парсинга (можно добавить любые)
REGIONS = [
    "Бишкек, Кыргызстан",
    "Ош, Кыргызстан",
]

# Ключевые слова поиска
SEARCH_KEYWORDS = [
    "салон красоты",
    "парикмахерская",
    "nail studio",
    "beauty salon",
]

# Сервисные категории — для определения услуги по тексту отзыва
SERVICE_KEYWORDS = {
    "Маникюр / Педикюр": [
        "маникюр", "педикюр", "ногт", "гель", "shellac", "шеллак",
        "покрытие", "nail", "наращивание ногтей",
    ],
    "Стрижка / Укладка": [
        "стрижк", "укладк", "причёск", "haircut", "укоротила", "волосы",
        "стриглась", "локоны", "блоу", "blow dry",
    ],
    "Окрашивание": [
        "окрашивани", "краска", "цвет", "highlight", "блонд", "мелирование",
        "балаяж", "тонирование", "color",
    ],
    "Брови / Ресницы": [
        "бров", "ресниц", "архитектура бровей", "ламинирование",
        "наращивание ресниц", "lashes", "brows",
    ],
    "Макияж": [
        "макияж", "визаж", "makeup", "свадебный", "вечерний",
    ],
    "Массаж / SPA": [
        "массаж", "spa", "спа", "релакс", "антицеллюлит",
    ],
    "Эпиляция": [
        "эпиляция", "воск", "шугаринг", "depilation", "лазер",
    ],
    "Уход за лицом": [
        "чистка лица", "пилинг", "уход за кожей", "facial", "карбокситерапия",
    ],
}

MAX_REVIEWS_PER_SALON = 20   # отзывов на салон (бесплатно до 5 без Details)
output_file = "beauty_salons_report.xlsx"

# ─── GOOGLE MAPS API ──────────────────────────────────────────────────────────

BASE_SEARCH  = "https://maps.googleapis.com/maps/api/place/textsearch/json"
BASE_DETAILS = "https://maps.googleapis.com/maps/api/place/details/json"


def search_places(query: str, api_key: str) -> list[dict]:
    """Поиск мест через Text Search API (постранично)."""
    results, token = [], None
    while True:
        params = {"query": query, "key": api_key, "language": "ru"}
        if token:
            params["pagetoken"] = token
            print("   [Google API] Ждем 2 секунды перед следующей страницей поиска...")
            time.sleep(2)          # Google требует паузу перед следующей страницей
        
        print(f"   [Google API] Выполняем поиск: {query}")
        r = requests.get(BASE_SEARCH, params=params, timeout=15)
        data = r.json()
        if data.get("status") not in ("OK", "ZERO_RESULTS"):
            error_msg = data.get('error_message', '')
            status_code = data.get('status')
            print(f"  [Ошибка]  Search error: {status_code} - {error_msg}")
            if status_code == "REQUEST_DENIED":
                raise ValueError(f"Ошибка API ключа (REQUEST_DENIED): {error_msg}")
            break
        results.extend(data.get("results", []))
        token = data.get("next_page_token")
        if not token:
            break
    return results


def get_place_details(place_id: str, api_key: str) -> dict:
    """Детальная информация + отзывы через Place Details API."""
    params = {
        "place_id": place_id,
        "fields": (
            "name,formatted_address,formatted_phone_number,"
            "website,rating,user_ratings_total,"
            "opening_hours,reviews,types,url"
        ),
        "language": "ru",
        "key": api_key,
    }
    print(f"   [Google API] Запрос деталей для: {place_id}")
    r = requests.get(BASE_DETAILS, params=params, timeout=15)
    data = r.json()
    if data.get("status") != "OK":
        return {}
    return data.get("result", {})


# ─── АНАЛИЗ ОТЗЫВОВ ──────────────────────────────────────────────────────────

def detect_service(text: str) -> str:
    """Определить услугу по тексту отзыва."""
    text_lower = text.lower()
    for service, keywords in SERVICE_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            return service
    return "Другое / Общий"


def relative_time_to_date(relative_time_description: str) -> str:
    """Приблизительный перевод 'X months ago' → период."""
    s = relative_time_description.lower()
    now = datetime.now()
    if "day" in s or "день" in s or "дня" in s or "дней" in s:
        nums = re.findall(r"\d+", s)
        days = int(nums[0]) if nums else 1
        return (now.replace(day=max(1, now.day - days))).strftime("%Y-%m")
    if "week" in s or "недел" in s:
        nums = re.findall(r"\d+", s)
        weeks = int(nums[0]) if nums else 1
        return (now.replace(day=max(1, now.day - weeks * 7))).strftime("%Y-%m")
    if "month" in s or "месяц" in s or "месяца" in s or "месяцев" in s:
        nums = re.findall(r"\d+", s)
        months = int(nums[0]) if nums else 1
        m = now.month - months
        y = now.year + m // 12
        m = m % 12 or 12
        return f"{y}-{m:02d}"
    if "year" in s or "год" in s or "года" in s or "лет" in s:
        nums = re.findall(r"\d+", s)
        years = int(nums[0]) if nums else 1
        return f"{now.year - years}"
    return relative_time_description


def analyze_reviews(reviews: list[dict]) -> dict:
    """Агрегированная аналитика по отзывам."""
    service_counter   = Counter()
    period_counter    = Counter()
    rating_counter    = Counter()
    reviewer_freq     = Counter()

    for rv in reviews:
        text   = rv.get("text", "")
        rating = rv.get("rating", 0)
        period = relative_time_to_date(rv.get("relative_time_description", ""))
        author = rv.get("author_name", "Аноним")

        service_counter[detect_service(text)] += 1
        period_counter[period] += 1
        rating_counter[rating] += 1
        reviewer_freq[author]  += 1

    return {
        "service_counter": service_counter,
        "period_counter":  period_counter,
        "rating_counter":  rating_counter,
        "reviewer_freq":   reviewer_freq,
    }


# ─── EXCEL ОТЧЁТ ─────────────────────────────────────────────────────────────

# Цвета
C_HEADER    = "1F3864"   # тёмно-синий
C_SUBHEADER = "2E75B6"   # средне-синий
C_ACCENT    = "D6E4F0"   # светло-голубой
C_GREEN     = "E2EFDA"
C_YELLOW    = "FFF2CC"
C_WHITE     = "FFFFFF"
C_ORANGE    = "FCE4D6"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def hdr(ws, row, col, value, bg=C_HEADER, fg="FFFFFF", bold=True, wrap=False, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border    = THIN_BORDER
    return c


def cell(ws, row, col, value, bg=C_WHITE, bold=False, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, name="Arial", size=10)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = THIN_BORDER
    return c


def star_rating(rating: float) -> str:
    full = int(rating)
    half = 1 if (rating - full) >= 0.5 else 0
    return "★" * full + ("½" if half else "") + f"  ({rating:.1f})"


# ── Лист 1: Сводка по салонам ─────────────────────────────────────────────────
def sheet_salons(wb, salons_data, regions):
    ws = wb.create_sheet("📋 Салоны")
    ws.sheet_view.showGridLines = False

    # Заголовок
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value     = "🌸  Отчёт: Салоны красоты — Google Maps"
    t.font      = Font(bold=True, size=16, color="FFFFFF", name="Arial")
    t.fill      = PatternFill("solid", start_color=C_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:L2")
    ws["A2"].value     = f"Дата выгрузки: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Регионы: {', '.join(regions)}"
    ws["A2"].font      = Font(italic=True, size=10, name="Arial", color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill      = PatternFill("solid", start_color="EBF3FB")

    cols = [
        "№", "Название", "Регион", "Адрес", "Телефон", "Сайт",
        "Рейтинг", "Кол-во отзывов", "Топ-услуга", "Часы работы",
        "Ссылка на Google Maps", "Тип заведения",
    ]
    widths = [4, 30, 18, 35, 18, 25, 12, 14, 22, 30, 35, 20]

    for i, (col_name, w) in enumerate(zip(cols, widths), 1):
        hdr(ws, 3, i, col_name, bg=C_SUBHEADER)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22

    for idx, sd in enumerate(salons_data, 1):
        r   = idx + 3
        bg  = C_ACCENT if idx % 2 == 0 else C_WHITE
        top_service = sd["analysis"]["service_counter"].most_common(1)
        top_service = top_service[0][0] if top_service else "—"
        hours = ""
        if sd.get("opening_hours"):
            periods = sd["opening_hours"].get("weekday_text", [])
            hours   = "\n".join(periods[:3]) + ("…" if len(periods) > 3 else "")

        cell(ws, r, 1,  idx,                  bg=bg, align="center")
        cell(ws, r, 2,  sd.get("name","—"),   bg=bg, bold=True)
        cell(ws, r, 3,  sd.get("region","—"), bg=bg)
        cell(ws, r, 4,  sd.get("address","—"),bg=bg)
        cell(ws, r, 5,  sd.get("phone","—"),  bg=bg)
        cell(ws, r, 6,  sd.get("website","—"),bg=bg)
        cell(ws, r, 7,  star_rating(sd.get("rating",0)), bg=bg, align="center")
        cell(ws, r, 8,  sd.get("total_ratings",0), bg=bg, align="center")
        cell(ws, r, 9,  top_service,           bg=bg)
        cell(ws, r, 10, hours,                 bg=bg)
        cell(ws, r, 11, sd.get("maps_url","—"),bg=bg)
        cell(ws, r, 12, ", ".join(sd.get("types",[])[:2]), bg=bg)
        ws.row_dimensions[r].height = 38

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{len(salons_data)+3}"
    return ws


# ── Лист 2: Все отзывы ────────────────────────────────────────────────────────
def sheet_reviews(wb, salons_data):
    ws = wb.create_sheet("💬 Отзывы")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"].value     = "💬  Все отзывы"
    ws["A1"].font      = Font(bold=True, size=15, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Салон", "Регион", "Автор", "Рейтинг", "Период", "Услуга (авто)", "Текст отзыва", "Язык?", "Тональность"]
    widths = [28, 18, 22, 10, 12, 24, 60, 10, 14]
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        hdr(ws, 2, i, c, bg=C_SUBHEADER)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 20

    row = 3
    for sd in salons_data:
        for rv in sd.get("reviews", []):
            rating  = rv.get("rating", 0)
            text    = rv.get("text", "")
            service = detect_service(text)
            period  = relative_time_to_date(rv.get("relative_time_description",""))
            # Простая тональность по рейтингу
            if rating >= 4:
                sentiment, s_bg = "Позитивный 😊", C_GREEN
            elif rating == 3:
                sentiment, s_bg = "Нейтральный 😐", C_YELLOW
            else:
                sentiment, s_bg = "Негативный 😟", C_ORANGE

            bg = s_bg
            cell(ws, row, 1, sd.get("name",""),   bg=bg)
            cell(ws, row, 2, sd.get("region",""),  bg=bg)
            cell(ws, row, 3, rv.get("author_name","Аноним"), bg=bg)
            cell(ws, row, 4, "★" * rating + f" ({rating})", bg=bg, align="center")
            cell(ws, row, 5, period,               bg=bg, align="center")
            cell(ws, row, 6, service,              bg=bg)
            cell(ws, row, 7, text,                 bg=bg, wrap=True)
            # Определяем язык приблизительно (кириллица vs латиница)
            lang = "RU/KG" if re.search(r"[а-яё]", text, re.I) else "EN/Other"
            cell(ws, row, 8, lang,                 bg=bg, align="center")
            cell(ws, row, 9, sentiment,            bg=bg, align="center")
            ws.row_dimensions[row].height = 50
            row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:I{row-1}"
    return ws


# ── Лист 3: Аналитика услуг ──────────────────────────────────────────────────
def sheet_services(wb, salons_data):
    ws = wb.create_sheet("💅 Услуги")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value     = "💅  Распределение упоминаний услуг в отзывах"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color=C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Суммарный счётчик
    total_services: Counter = Counter()
    for sd in salons_data:
        total_services += sd["analysis"]["service_counter"]

    for c, w in zip("ABCDE", [30, 16, 16, 22, 30]):
        ws.column_dimensions[c].width = w

    for col, label in enumerate(["Услуга", "Упоминаний", "% от общего", "Рейтинг услуги (ср)", "Комментарий"], 1):
        hdr(ws, 2, col, label, bg=C_SUBHEADER)

    total_cnt = sum(total_services.values()) or 1
    row = 3
    for svc, cnt in total_services.most_common():
        pct = cnt / total_cnt
        # Средний рейтинг для услуги
        ratings_for_svc = []
        for sd in salons_data:
            for rv in sd.get("reviews", []):
                if detect_service(rv.get("text","")) == svc:
                    ratings_for_svc.append(rv.get("rating", 0))
        avg_r = sum(ratings_for_svc) / len(ratings_for_svc) if ratings_for_svc else 0

        bg = C_ACCENT if row % 2 == 0 else C_WHITE
        cell(ws, row, 1, svc,         bg=bg, bold=True)
        cell(ws, row, 2, cnt,         bg=bg, align="center")
        ws.cell(row=row, column=3).value       = pct
        ws.cell(row=row, column=3).number_format = "0.0%"
        ws.cell(row=row, column=3).fill        = PatternFill("solid", start_color=bg)
        ws.cell(row=row, column=3).border      = THIN_BORDER
        ws.cell(row=row, column=3).alignment   = Alignment(horizontal="center")
        cell(ws, row, 4, f"{avg_r:.1f} ★" if avg_r else "—", bg=bg, align="center")
        # Авто-комментарий
        if avg_r >= 4.5:
            comment = "Отличные отзывы 🟢"
        elif avg_r >= 3.5:
            comment = "Хорошие отзывы 🟡"
        elif avg_r > 0:
            comment = "Есть жалобы 🔴"
        else:
            comment = "Нет рейтинга"
        cell(ws, row, 5, comment, bg=bg)
        row += 1

    # Мини-диаграмма (bar chart)
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Упоминания услуг"
    chart.y_axis.title = "Кол-во"
    chart.x_axis.title = "Услуга"
    chart.style = 10
    chart.width = 30
    chart.height = 16

    data_ref = Reference(ws, min_col=2, min_row=2, max_row=row-1)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "G2")
    return ws


# ── Лист 4: Активность по периодам ───────────────────────────────────────────
def sheet_activity(wb, salons_data):
    ws = wb.create_sheet("📅 Активность")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"].value     = "📅  Суммарная активность отзывов по всем салонам"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color=C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    period_all: Counter = Counter()
    for sd in salons_data:
        period_all += sd["analysis"]["period_counter"]

    for c, w in zip("ABCD", [20, 18, 20, 28]):
        ws.column_dimensions[c].width = w
    for col, label in enumerate(["Период", "Кол-во отзывов", "% от общего", "Тренд"], 1):
        hdr(ws, 2, col, label, bg=C_SUBHEADER)

    total = sum(period_all.values()) or 1
    sorted_periods = sorted(period_all.items())
    prev = 0
    for r, (period, cnt) in enumerate(sorted_periods, 3):
        trend = ("↑" if cnt > prev else ("↓" if cnt < prev and prev > 0 else "→")) if prev else "—"
        bg    = C_GREEN if cnt >= (total / len(sorted_periods)) else C_WHITE
        cell(ws, r, 1, period, bg=bg, align="center")
        cell(ws, r, 2, cnt,    bg=bg, align="center")
        ws.cell(row=r, column=3).value        = cnt / total
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=3).fill         = PatternFill("solid", start_color=bg)
        ws.cell(row=r, column=3).border       = THIN_BORDER
        ws.cell(row=r, column=3).alignment    = Alignment(horizontal="center")
        cell(ws, r, 4, trend, bg=bg, align="center")
        prev = cnt
    return ws


# ── Лист 5: Постоянные клиенты ───────────────────────────────────────────────
def sheet_loyal(wb, salons_data):
    ws = wb.create_sheet("🏆 Клиенты")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value     = "🏆  Авторы с несколькими отзывами (лояльные клиенты)"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color=C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    freq_all: Counter = Counter()
    author_reviews = defaultdict(list)
    for sd in salons_data:
        salon_name = sd.get("name", "—")
        freq_all += sd["analysis"]["reviewer_freq"]
        for rv in sd.get("reviews", []):
            author = rv.get("author_name", "Аноним")
            text = rv.get("text", "").strip()
            if text:
                author_reviews[author].append({"salon": salon_name, "text": text})

    for c, w in zip("ABCDE", [35, 18, 20, 25, 75]):
        ws.column_dimensions[c].width = w
    for col, label in enumerate(["Автор", "Кол-во отзывов", "Статус", "Примечание", "Оставленные отзывы"], 1):
        hdr(ws, 2, col, label, bg=C_SUBHEADER)

    row = 3
    for author, cnt in freq_all.most_common():
        if cnt < 2:
            continue
        if cnt >= 5:
            status, bg = "⭐ Суперклиент", C_GREEN
        elif cnt >= 3:
            status, bg = "🔁 Постоянный",  C_ACCENT
        else:
            status, bg = "✔️ Лояльный",    C_WHITE
        
        note = f"Оставил(а) {cnt} отзыва — ценный голос"
        
        revs = author_reviews.get(author, [])
        if revs:
            reviews_text = "\n\n".join([f"📍 {r['salon']}:\n💬 \"{r['text']}\"" for r in revs])
        else:
            reviews_text = "Отзывы без текста (только рейтинг)"

        cell(ws, row, 1, author, bg=bg, bold=True)
        cell(ws, row, 2, cnt,    bg=bg, align="center")
        cell(ws, row, 3, status, bg=bg, align="center")
        cell(ws, row, 4, note,   bg=bg)
        cell(ws, row, 5, reviews_text, bg=bg, wrap=True)
        
        # Увеличим высоту строки для удобного чтения отзывов
        ws.row_dimensions[row].height = min(150, max(40, len(revs) * 35))
        row += 1

    if row == 3:
        ws.merge_cells("A3:E3")
        ws["A3"].value = "Все авторы оставили по одному отзыву — нет повторных."
        ws["A3"].alignment = Alignment(horizontal="center")
    return ws


# ── Лист 6: Сводная статистика ────────────────────────────────────────────────
def sheet_summary(wb, salons_data, regions):
    ws = wb.create_sheet("📊 Итог", 0)   # первый лист
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

    ws.merge_cells("A1:B1")
    ws["A1"].value     = "📊  Итоговая статистика"
    ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color=C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    all_reviews = [rv for sd in salons_data for rv in sd.get("reviews", [])]
    all_ratings = [rv.get("rating", 0) for rv in all_reviews]
    avg_rating  = sum(all_ratings) / len(all_ratings) if all_ratings else 0

    valid_salons = [s for s in salons_data if s.get("total_ratings", 0) >= 5]
    top20 = sorted(valid_salons, key=lambda x: (x.get("rating", 0), x.get("total_ratings", 0)), reverse=True)[:20]
    
    service_all: Counter = Counter()
    for sd in salons_data:
        service_all += sd["analysis"]["service_counter"]
    top_service = service_all.most_common(1)[0][0] if service_all else "—"

    stats = [
        ("", ""),
        ("🏪  Всего салонов найдено",           len(salons_data)),
        ("💬  Всего собрано отзывов",            len(all_reviews)),
        ("⭐  Средний рейтинг по всем салонам",  f"{avg_rating:.2f}"),
        ("💅  Самая упоминаемая услуга",          top_service),
        ("🌍  Регионы охвата",                   ", ".join(regions)),
        ("📅  Дата парсинга",                    datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("", ""),
        ("🏆  ТОП-20 салонов (мин. 5 отзывов)", ""),
    ]
    for sd in top20:
        stats.append((f"   • {sd['name']}", f"{sd.get('rating','—')} ★ ({sd.get('total_ratings',0)} отзывов)"))

    for r, (k, v) in enumerate(stats, 2):
        if k == "":
            ws.row_dimensions[r].height = 8
            continue
        c1 = ws.cell(row=r, column=1, value=k)
        c2 = ws.cell(row=r, column=2, value=v)
        for c in [c1, c2]:
            c.font      = Font(name="Arial", size=11)
            c.alignment = Alignment(vertical="center")
            c.border    = THIN_BORDER
        c1.font = Font(name="Arial", size=11, bold=True)
        if "ТОП" in k:
            c1.fill = PatternFill("solid", start_color=C_YELLOW)
            c2.fill = PatternFill("solid", start_color=C_YELLOW)
        elif k.startswith("   •"):
            c1.fill = PatternFill("solid", start_color=C_GREEN)
            c2.fill = PatternFill("solid", start_color=C_GREEN)
        else:
            c1.fill = PatternFill("solid", start_color=C_ACCENT)
            c2.fill = PatternFill("solid", start_color=C_WHITE)
        ws.row_dimensions[r].height = 22
    return ws


# ─── ОСНОВНОЙ PIPELINE ────────────────────────────────────────────────────────

def run(api_key, regions, keywords, output_file, max_reviews=20, status_callback=None):
    print("=" * 60)
    print("  Beauty Salon Scraper - Google Maps -> Excel")
    print("=" * 60)

    salons_data = []
    seen_ids    = set()

    for region in regions:
        for keyword in keywords:
            query = f"{keyword} {region}"
            print(f"\n[Поиск]: {query}")
            places = search_places(query, api_key)
            print(f"   Найдено мест: {len(places)}")

            for place in places:
                pid = place.get("place_id")
                if not pid or pid in seen_ids:
                    continue
                seen_ids.add(pid)

                print(f"   [Место] {place.get('name')} - получаю детали...")
                details = get_place_details(pid, api_key)
                if not details:
                    print(f"   [Пропуск] Детали не получены для {place.get('name')}")
                    continue

                reviews   = details.get("reviews", [])[:max_reviews]
                print(f"   [Данные] Получено отзывов: {len(reviews)}")
                analysis  = analyze_reviews(reviews)

                salons_data.append({
                    "name":          details.get("name", place.get("name", "—")),
                    "region":        region,
                    "address":       details.get("formatted_address", "—"),
                    "phone":         details.get("formatted_phone_number", "—"),
                    "website":       details.get("website", "—"),
                    "rating":        details.get("rating", 0),
                    "total_ratings": details.get("user_ratings_total", 0),
                    "opening_hours": details.get("opening_hours", {}),
                    "types":         details.get("types", []),
                    "maps_url":      details.get("url", "—"),
                    "reviews":       reviews,
                    "analysis":      analysis,
                })
                print(f"   [Успех] Салон {place.get('name')} добавлен. Всего собрано: {len(salons_data)}")
                if status_callback:
                    status_callback(f"Собран: {place.get('name')} (Всего: {len(salons_data)})")
                time.sleep(0.3)   # бережём квоту API

    if not salons_data:
        print("\n[Внимание]  Данные не получены. Проверьте API_KEY и регионы.")
        raise ValueError("По вашему запросу не найдено ни одного салона. Попробуйте изменить ключевые слова или регионы.")

    print(f"\n[Готово]  Собрано {len(salons_data)} уникальных салонов.")
    print("[Отчет] Создаю Excel отчёт...")

    wb = openpyxl.Workbook()
    # Удалить дефолтный лист
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheet_summary(wb, salons_data, regions)
    sheet_salons(wb, salons_data, regions)
    sheet_reviews(wb, salons_data)
    sheet_services(wb, salons_data)
    sheet_activity(wb, salons_data)
    sheet_loyal(wb, salons_data)

    wb.save(output_file)
    print(f"\n[Успех]  Готово! Файл сохранён: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    run()

